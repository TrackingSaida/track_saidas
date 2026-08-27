"""Exportação da tela de Registros (CSV / XLSX).

Reaproveita o filtro de GET /saidas/listar. Não altera o contrato da listagem.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

MAX_EXPORTAR_LIMIT = 10000
FORMATOS_EXPORTAR = ("csv", "xlsx")
OPERACAO_TZ = ZoneInfo("America/Sao_Paulo")

EXPORT_COLUMNS: List[Dict[str, str]] = [
    {"key": "codigo", "label": "Código"},
    {"key": "servico", "label": "Serviço"},
    {"key": "status", "label": "Status"},
    {"key": "acao", "label": "Última ação"},
    {"key": "entregador", "label": "Motoboy"},
    {"key": "data_hora_acao", "label": "Horário da ação"},
    {"key": "timestamp", "label": "Entrada no sistema"},
    {"key": "executado_por", "label": "Executado por"},
    {"key": "base", "label": "Base"},
    {"key": "is_grande", "label": "G"},
]

EXPORT_COLUMN_KEYS = tuple(col["key"] for col in EXPORT_COLUMNS)
EXPORT_COLUMN_LABELS = {col["key"]: col["label"] for col in EXPORT_COLUMNS}

MSG_SEM_REGISTROS = "Não há registros para exportar com os filtros atuais."
MSG_TETO = (
    "Há mais de 10.000 registros com os filtros atuais. "
    "Restrinja o período ou os filtros para exportar."
)
MSG_SEM_COLUNAS = "Selecione pelo menos uma coluna."
MSG_FORMATO = "Formato inválido. Use csv ou xlsx."

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExportarSaidasError(ValueError):
    def __init__(self, message: str):
        super().__init__(message)
        self.message = message


def resolver_formato(formato: Optional[str]) -> str:
    value = (formato or "").strip().lower()
    if value not in FORMATOS_EXPORTAR:
        raise ExportarSaidasError(MSG_FORMATO)
    return value


def resolver_colunas(colunas: Optional[Sequence[str]]) -> List[str]:
    if colunas is None:
        return list(EXPORT_COLUMN_KEYS)
    wanted: List[str] = []
    seen = set()
    for raw in colunas:
        key = str(raw or "").strip()
        if key not in EXPORT_COLUMN_LABELS or key in seen:
            continue
        wanted.append(key)
        seen.add(key)
    if not wanted:
        raise ExportarSaidasError(MSG_SEM_COLUNAS)
    return wanted


def formatar_status_export(status: Any) -> str:
    if status is None or str(status).strip() == "":
        return "—"
    s = str(status).replace("_", " ").strip()
    lower = s.lower()
    if lower == "na base":
        return "Na Base"
    if lower in {"saiu", "saiu para entrega"}:
        return "SAIU PARA ENTREGA"
    if lower in {"encerrado sistema", "encerrado pelo sistema", "encerrado"}:
        return "Encerrado"
    return s.upper()


def formatar_datetime_export(value: Any) -> str:
    if value is None or value == "":
        return ""
    dt: Optional[datetime] = None
    if isinstance(value, datetime):
        dt = value
    else:
        raw = str(value).strip()
        if not raw:
            return ""
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return raw
    if dt.tzinfo is not None:
        dt = dt.astimezone(OPERACAO_TZ).replace(tzinfo=None)
    return dt.strftime("%d/%m/%Y %H:%M")


def formatar_celula(key: str, item: Dict[str, Any]) -> str:
    if key == "status":
        return formatar_status_export(item.get("status"))
    if key == "is_grande":
        return "Sim" if item.get("is_grande") else "Não"
    if key in {"timestamp", "data_hora_acao"}:
        return formatar_datetime_export(item.get(key))
    if key == "acao":
        value = str(item.get("acao") or "").strip()
        return value or "Sem ação"
    value = item.get(key)
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"—", "-"}:
        return text
    return text


def montar_linhas(items: Iterable[Dict[str, Any]], colunas: Sequence[str]) -> List[List[str]]:
    rows: List[List[str]] = []
    for item in items:
        rows.append([formatar_celula(key, item) for key in colunas])
    return rows


def cabecalhos(colunas: Sequence[str]) -> List[str]:
    return [EXPORT_COLUMN_LABELS[key] for key in colunas]


def nome_arquivo_export(formato: str, agora: Optional[datetime] = None) -> str:
    momento = agora or datetime.now(OPERACAO_TZ)
    local = momento.astimezone(OPERACAO_TZ) if momento.tzinfo else momento
    stamp = local.strftime("%Y-%m-%d_%H%M")
    ext = "xlsx" if formato == "xlsx" else "csv"
    return f"registros_{stamp}.{ext}"


def gerar_csv(colunas: Sequence[str], items: Sequence[Dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    writer.writerow(cabecalhos(colunas))
    for linha in montar_linhas(items, colunas):
        writer.writerow(linha)
    return buf.getvalue().encode("utf-8-sig")


def gerar_xlsx(colunas: Sequence[str], items: Sequence[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"
    headers = cabecalhos(colunas)
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    text_keys = {"codigo"}
    for linha, item in zip(montar_linhas(items, colunas), items):
        ws.append(linha)
        row_idx = ws.max_row
        for col_idx, key in enumerate(colunas, start=1):
            if key in text_keys:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = "@"
                cell.value = str(item.get("codigo") or linha[col_idx - 1] or "")

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[letter].width = max(12, max_len + 2)

    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def gerar_arquivo(formato: str, colunas: Sequence[str], items: Sequence[Dict[str, Any]]) -> Tuple[bytes, str, str]:
    fmt = resolver_formato(formato)
    cols = resolver_colunas(colunas)
    filename = nome_arquivo_export(fmt)
    if fmt == "xlsx":
        return gerar_xlsx(cols, items), filename, XLSX_MEDIA_TYPE
    return gerar_csv(cols, items), filename, CSV_MEDIA_TYPE


def validar_total_exportacao(total: int) -> None:
    if int(total or 0) <= 0:
        raise ExportarSaidasError(MSG_SEM_REGISTROS)
    if int(total) > MAX_EXPORTAR_LIMIT:
        raise ExportarSaidasError(MSG_TETO)


def content_disposition(filename: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", filename or "registros.csv")
    return f'attachment; filename="{safe}"'
